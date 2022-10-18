from subprocess import Popen, PIPE
from difflib import SequenceMatcher
import os
import json
import pandas as pd
from openpyxl import Workbook, load_workbook
import re
import sys
import glob

from automate_utils import files
from automate_utils import scoring

def run_python_script(script, my_input, timeout=4, verbose=0):
    p = Popen(f'python "{script}"', stdin=PIPE, stdout=PIPE, shell=False)
    res= ''

    try:
        out = p.communicate(input=my_input.encode(), timeout=timeout)[0]
    except:
        if verbose>=3:
            print('Timed out...')
    else:
        res= out[:-2].decode()


    return res

def resave_excel_with_formula_final_score(excel_path, len_correct_script, bonus_task_from=None, start_col= 'D'):
    workbook = load_workbook(excel_path)
    curr_sheet = workbook[workbook.sheetnames[0]]

    end_col = chr(ord(start_col) + len_correct_script - 1)
    bonus_start_col = ''
    bonus_end_col = ''
    temp_len_correct_script= len_correct_script

    if bonus_task_from is not None:
        len_correct_script = (len_correct_script - (len_correct_script - bonus_task_from) - 1)
        end_col = chr(ord(start_col) + len_correct_script - 1)
        bonus_start_col = chr(ord(end_col) + 1)
        bonus_end_col = chr(ord(bonus_start_col) + (temp_len_correct_script - len_correct_script) - 1)

    for idx, i in enumerate(curr_sheet.rows):
        if idx < 1: continue

        if bonus_task_from is not None:
            if bonus_task_from<=1:
                final_score_formula = f'=VALUE(FIXED(SUM({start_col}{idx + 1}:{bonus_end_col}{idx + 1}), 3))'
            else:
                final_score_formula = f'=FIXED((VALUE(SUM({start_col}{idx + 1}:{end_col}{idx + 1})/{len_correct_script})+SUM({bonus_start_col}{idx + 1}:{bonus_end_col}{idx + 1})), 3)'

            i[temp_len_correct_script + 3].value = final_score_formula
        else:
            final_score_formula = f'=FIXED(VALUE(SUM({start_col}{idx + 1}:{end_col}{idx + 1})/{len_correct_script}), 3)'
            i[len_correct_script + 3].value = final_score_formula

    workbook.save(excel_path)


try:
    config_filename= sys.argv[1].strip()
except:
    config_filename= 'automate_config.json'


json_file = open(config_filename)
variables= json.load(json_file)

base_dir= variables['base_dir']
test_set_path= variables['test_set_path']
test_set_list_reference= variables['test_set_list_reference']
test_set_list_reference_delimiter= variables['test_set_list_reference_delimiter']
correct_script_set_path= variables['correct_script_set_path']
correct_script_set= variables['correct_script_set']
test_set_file_prefix= variables['test_set_file_prefix']
test_set_file_suffix= variables['test_set_file_suffix']
nim_input_type= variables['nim_input_type']
replace_zero_nim= variables['replace_zero_nim']
replace_zero_nim_with= variables['replace_zero_nim_with']
input_set= variables['input_set']
scoring_format_type_list= variables['scoring_format_type_list']
numeric_score_weight= variables['numeric_score_weight']
caption_score_weight= variables['caption_score_weight']
bonus_task_from= variables['bonus_task_from']
bonus_task_score= variables['bonus_task_score']
add_extra_score= variables['add_extra_score']
extra_score_list= variables['extra_score_list']
save_output= variables['save_output']
write_file_mode= variables['write_file_mode']
input_file_name= variables['input_file_name']
output_file_name= variables['output_file_name']
output_file_format= variables['output_file_format']
new_column_name_prefix= variables['new_column_name_prefix']
new_column_name_suffix= variables['new_column_name_suffix']
verbose= variables['verbose']


json_file.close()

if write_file_mode=='w':
    try:
        data= pd.read_csv(test_set_list_reference, delimiter=test_set_list_reference_delimiter)
    except:
        data= pd.read_csv(os.path.join(base_dir, test_set_list_reference), delimiter=test_set_list_reference_delimiter)

elif write_file_mode=='a':
    if input_file_name.split('.')[-1]=='csv':
        data= pd.read_csv(os.path.join(base_dir, input_file_name))
    elif input_file_name.split('.')[-1] == 'xlsx':
        data = pd.read_excel(os.path.join(base_dir, input_file_name))
    else:
        data= pd.read_csv(os.path.join(base_dir, input_file_name))
else:
    try:
        data= pd.read_csv(test_set_list_reference, delimiter=test_set_list_reference_delimiter)
    except:
        data= pd.read_csv(os.path.join(base_dir, test_set_list_reference), delimiter=test_set_list_reference_delimiter)

def get_input_sets(input_set, nim=None):
    if input_set == 'nim':
        print(nim)
        if nim_input_type == 0:
            temp_input_set = list(nim[-2:])
        elif nim_input_type == 1:
            temp_input_set = list(nim[-3:])
        elif nim_input_type == 2:
            temp_input_set = nim[-2:]
        else:
            temp_input_set = list(nim[-2:])

        temp_input_set = [str(replace_zero_nim_with) if j == '0' else j for j in temp_input_set]

        if isinstance(temp_input_set, list):
            temp_input_set = '\n'.join(temp_input_set)

        temp_input_set = [temp_input_set]

    else:
        temp_input_set = ['\n'.join([str(k) for k in j]) for j in input_set]

    return temp_input_set


# print(data)

# H071191069
# Type 0: a=6; b=9
# Type 1: a=0; b=6; c=9;
# Type 2: a=69

# nim_input_type= 0

# If replace_zero_nim is True and replace_zero_nim_with==3
# H071191040; type 0
# a=4; b=3

# Scoring Format:
# 0 : caption and numeric
# 1 : numeric
# 2 : caption

# replace_zero_nim= True
# replace_zero_nim_with= 3
#
# numeric_score_weight = 0.85
# caption_score_weight = 0.15

# write_file_mode
# w: write and replace
# a: write and append

if __name__=='__main__':
    final_score_column_name= 'final_skor'

    if write_file_mode=='w':
        new_column_list = []

        for new_column in range(len(correct_script_set)):
            new_column_name= f'{new_column_name_prefix}_{new_column+1}'+('' if new_column_name_suffix=='' else '_'+new_column_name_suffix)
            new_column_list.append(new_column_name)

            data[new_column_name]= 0

        new_column_list.append(final_score_column_name)
        data[final_score_column_name]= 0

    elif write_file_mode=='a':
        new_column_list= [i for i in data.columns if (test_set_file_prefix in i and test_set_file_suffix in i and i!='no') or i=='final_skor']
    else:
        new_column_list = []

        for new_column in range(len(correct_script_set)):
            new_column_name= f'{new_column_name_prefix}_{new_column+1}'+('' if new_column_name_suffix=='' else '_'+new_column_name_suffix)
            new_column_list.append(new_column_name)

            data[new_column_name]= 0

        new_column_list.append(final_score_column_name)
        data[final_score_column_name]= 0

    import time
    start= time.time()

    # desired_output_list= [correct_script if correct_script is not None else None for correct_script in correct_script_set]
    # desired_output_list= {correct_script:[7] for correct_script in correct_script_set}
    desired_output_dict= {}
    desired_output_numeric_dict= {}
    desired_output_len_dict= {}

    for correct_script_idx in range(len(correct_script_set)):
        k= f'no{correct_script_idx+1}'
        desired_output_dict[k] = None
        desired_output_numeric_dict[k] = None
        desired_output_len_dict[k]= None

        if correct_script_set[correct_script_idx] is not None:
            correct_script_set[correct_script_idx]= os.path.join(base_dir, correct_script_set_path, correct_script_set[correct_script_idx])

            desired_output_dict[k] = []
            desired_output_numeric_dict[k] = []
            desired_output_len_temp = 0

            if input_set[correct_script_idx]!='nim':
                temp_input_set= get_input_sets(input_set[correct_script_idx])

                for idx, test_set in enumerate(temp_input_set):
                    desired_output_dict[k].append(run_python_script(correct_script_set[correct_script_idx], test_set, verbose=verbose))
                    desired_output_numeric_dict[k].append(list(map(float, scoring.get_numeric(desired_output_dict[k][idx]))))
                    desired_output_len_temp+= len(desired_output_numeric_dict[k][idx])

                desired_output_len_dict[k]= desired_output_len_temp


    for nim_idx, raw_nim in enumerate(os.listdir(os.path.join(base_dir, test_set_path))):
        expected_check_script_set = [os.path.join(test_set_path, raw_nim, f'{test_set_file_prefix}{i + 1}{test_set_file_suffix}.py') for i in range(len(correct_script_set))]
        actual_check_script_set= sorted(glob.glob(os.path.join(base_dir, test_set_path, raw_nim) + '/*.py'))
        check_script_set= files.get_similiar_filename(expected_check_script_set, actual_check_script_set, test_set_file_prefix, test_set_file_suffix)

        # if nim_idx>=5:
        #     break

        try:
            nim= re.search(r'H\d{9}', raw_nim.strip().upper()).group()
        except:
            if verbose>=1:
                print('NIM not found in folder:',raw_nim)
            continue

        # if nim not in ['H071221104']:
        #     continue

        try:
            row_index = data[data['nim'] == nim].index[0]
        except:
            if verbose>=1:
                print('NIM not found in sample list:',nim)
            continue

        if verbose>=2:
            print(f'NIM: {nim}')

        score_list= []

        for correct_script_idx in range(len(correct_script_set)):
            check_script = check_script_set[correct_script_idx]
            scoring_format_type= scoring_format_type_list[correct_script_idx] if isinstance(scoring_format_type_list, list) else scoring_format_type_list

            if check_script is None and input_set[correct_script_idx] is not None:
                score_list.append(0)
                continue

            if input_set[correct_script_idx] is None:
                score_list.append(100)
                continue

            correct_script_key= f'no{correct_script_idx + 1}'

            numeric_score_list = []
            caption_score_list = []

            if input_set[correct_script_idx] == 'nim':
                desired_output_dict[correct_script_key] = []
                desired_output_numeric_dict[correct_script_key] = []
                desired_output_len_temp = 0
                temp_input_set = get_input_sets(input_set[correct_script_idx], nim)

                for idx, test_set in enumerate(temp_input_set):
                    desired_output_dict[correct_script_key].append(
                        run_python_script(correct_script_set[correct_script_idx], test_set, verbose=verbose))
                    desired_output_numeric_dict[correct_script_key].append(
                        list(map(float, scoring.get_numeric(desired_output_dict[correct_script_key][idx]))))
                    desired_output_len_temp += len(desired_output_numeric_dict[correct_script_key][idx])

                desired_output_len_dict[correct_script_key] = desired_output_len_temp
            else:
                temp_input_set = get_input_sets(input_set[correct_script_idx])

            for input_set_idx, test_set in enumerate(temp_input_set):
                given_output= run_python_script(check_script, test_set, verbose=verbose)
                given_output_numeric= list(map(float, scoring.get_numeric(given_output)))

                # print(given_output)
                # print(given_output_numeric)
                caption_score_list.append(scoring.similar(desired_output_dict[correct_script_key][input_set_idx], given_output))

                given_output_offset= 0
                # print(f'd: {desired_output_numeric_dict[correct_script_key]}, g: {given_output_numeric}')
                for numeric_idx in range(len(given_output_numeric)):
                    try:
                        if float(desired_output_numeric_dict[correct_script_key][input_set_idx][numeric_idx]) == float(given_output_numeric[numeric_idx+given_output_offset]):
                            numeric_score_list.append(100)
                        elif abs(float(desired_output_numeric_dict[correct_script_key][input_set_idx][numeric_idx]) - float(given_output_numeric[numeric_idx+given_output_offset])) < 1:
                            numeric_score_list.append(90)
                        else:
                            if len(desired_output_numeric_dict[correct_script_key])+given_output_offset<len(given_output_numeric):
                                given_output_offset+= 1

                                if float(desired_output_numeric_dict[correct_script_key][input_set_idx][numeric_idx]) == float(given_output_numeric[numeric_idx + given_output_offset]):
                                    numeric_score_list.append(100)
                                elif abs(float(desired_output_numeric_dict[correct_script_key][input_set_idx][numeric_idx]) - float(given_output_numeric[numeric_idx + given_output_offset])) < 1:
                                    numeric_score_list.append(90)
                                else:
                                    numeric_score_list.append(0)
                                    given_output_offset-= 1

                    except:
                        break

            print('Given',given_output_numeric)
            print('Desired',desired_output_numeric_dict)
            if len(numeric_score_list)<desired_output_len_dict[correct_script_key]:
                #HAPUS NANTI
                print(f'Desired num output: {desired_output_len_dict[correct_script_key]}')
                print(f'Given num output: {len(numeric_score_list)}')
                temp= len(numeric_score_list)
                a= [0 for _ in range(desired_output_len_dict[correct_script_key]-temp)]
                # print(a)
                numeric_score_list.extend(a)

            if verbose>=2:
                print(f'Numeric Score List: {numeric_score_list}')
                print(f'Caption Score List: {caption_score_list}')
                print()

            numeric_score= 0 if len(numeric_score_list)<=0 else sum(numeric_score_list)/len(numeric_score_list)
            caption_score= (0 if len(caption_score_list)<=0 else sum(caption_score_list)/len(caption_score_list))*100

            if scoring_format_type==0:
                total_curr_number_score= (numeric_score*numeric_score_weight) + (caption_score*caption_score_weight)
            elif scoring_format_type==1:
                total_curr_number_score= numeric_score
            elif scoring_format_type==2:
                total_curr_number_score= caption_score
            else:
                total_curr_number_score= (numeric_score*numeric_score_weight) + (caption_score*caption_score_weight)

            if add_extra_score:
                total_curr_number_score+= extra_score_list[correct_script_idx]
                total_curr_number_score= 100 if total_curr_number_score>=100 else total_curr_number_score

            if bonus_task_from is not None and correct_script_idx+1>=bonus_task_from:
                # print('------------------------------------')
                # print('no: ',correct_script_idx+1)
                # print('bobot: ', bonus_task_score[correct_script_idx+1-bonus_task_from])
                # print('------------------------------------')
                # print(6/0)
                # total_curr_number_score*= (bonus_task_score/100)
                bts= bonus_task_score[correct_script_idx+1-bonus_task_from] if isinstance(bonus_task_score, list) else bonus_task_score
                total_curr_number_score= total_curr_number_score/100*bts

            score_list.append(total_curr_number_score)

            if verbose==4:
                print(f'Total Numeric Score of No. {correct_script_idx+1}: {numeric_score}')
                print(f'Total Caption Score of No. {correct_script_idx+1}: {caption_score}')
                print(f'Final Score of No. {correct_script_idx+1}: {total_curr_number_score}')
                print('\n----------------------------------------\n')

            if input_set[correct_script_idx] == 'nim':
                desired_output_dict[correct_script_key] = []
                desired_output_numeric_dict[correct_script_key] = []
                desired_output_len_temp = 0

        for idx, score in enumerate(score_list):
            if score>float(data.loc[row_index, new_column_list[idx]]):
                data.loc[row_index, new_column_list[idx]] = float(f'{score:.3f}')

        final_score= 0

        if bonus_task_from is not None:
            for idx, j in enumerate([data.loc[row_index, i] for i in data.columns if test_set_file_prefix in i and test_set_file_suffix in i and i!='no']):
                if idx+1==bonus_task_from and bonus_task_from!=1:
                    final_score/= (len(correct_script_set)-(len(correct_script_set)-bonus_task_from)-1)

                final_score += j

        else:
            final_score= sum([data.loc[row_index, i] for i in data.columns if test_set_file_prefix in i and test_set_file_suffix in i and i!='no'])/len(correct_script_set)


        data.loc[row_index, new_column_list[-1]] = float(f'{final_score:.3f}')

        if verbose==1:
            print(f'Nim: {nim}, Score List: {score_list}, Final Score: {final_score:.3f}')


    if output_file_format=='csv':
        data.to_csv(os.path.join(base_dir, f'{output_file_name}.{output_file_format}'), index=False)
    elif output_file_format == 'excel':
        excel_path= os.path.join(base_dir, f'{output_file_name}.xlsx')

        data.T.reset_index().T.to_excel(excel_path, header=None, index=False)
        try:
            resave_excel_with_formula_final_score(excel_path, len(correct_script_set), bonus_task_from)
        except:
            if verbose==4:
                print('Something is wrong when trying to formula-ing the final score')

    else:
        data.to_csv(os.path.join(base_dir, f'{output_file_name}.csv'), index=False)

    print('Program selesai:',time.time()-start)